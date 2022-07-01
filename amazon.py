from selenium import webdriver
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

gorsel_yolu = "resimler"
arama = "bluetooth kulaklık"  # input("Aranacak ifadeyi yazın: ")

# excel oluştur veya varsa aç
excel_yolu = "excel/amazon_urunleri.xlsx"
if os.path.exists(excel_yolu):
    wb = load_workbook(excel_yolu)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # başlıkları ekle
    ws.append(["TARİH", "ID", "RESİM", "AD", "FİYAT"])

# tarayıcı oluştur
driver = "C:\\Users\\erenMustafaOzdal\\.wdm\\drivers\\geckodriver\\win64\\v0.31.0\\geckodriver.exe"
tarayici = webdriver.Firefox(executable_path=driver)
tarayici.maximize_window()
tarayici.get("https://www.amazon.com.tr")

# bekleme nesnesi oluştur
bekle = WebDriverWait(tarayici, 30)

# çerezi kabul et
bekle.until(ec.visibility_of_element_located(
    (By.ID, "sp-cc-accept")
)).click()

# arama yapalım
while True:
    try:
        tarayici.find_element(By.ID, "twotabsearchtextbox").send_keys(arama)
        tarayici.find_element(By.ID, "nav-search-submit-button").click()
        break
    except:
        print("Hata oluştu. 5 saniye sonra tekrar denenecek.")
        sleep(5)
        tarayici.refresh()


# sonraki sayfada ürünlerin gözükmesini bekle
def urunleri_tara():
    urunler = bekle.until(ec.visibility_of_all_elements_located(
        (By.XPATH, "//div[@data-component-type='s-search-result' and not(contains(@style, 'display: none'))]")
    ))

    print()
    print("#"*50)
    print("Ürün sayısı: ", len(urunler))
    print("#"*50)
    print()

    for urun in urunler:
        # id
        id = urun.get_attribute("data-asin")
        print("id:", id)

        # resimleri indir
        resim = f"resimler/{id}.png"
        urun.find_element(By.CLASS_NAME, "s-image").screenshot(resim)

        # ad
        ad = urun.find_element(By.TAG_NAME, "h2").text
        print("ad:", ad)

        # fiyat
        try:
            lira = int(urun.find_element(By.CLASS_NAME, "a-price-whole").text.replace(".", ""))
            kurus = int(urun.find_element(By.CLASS_NAME, "a-price-fraction").text)
            fiyat = float(f"{lira}.{kurus}")
        except:
            fiyat = "Fiyat yok"
        print("fiyat:", fiyat)


        print("-"*50)

        # excele bilgi ekle
        ws.append([
            datetime.now(),
            id,
            f'=HYPERLINK("{os.getcwd()}\\{resim}", "RESİM")',
            ad,
            fiyat
        ])


while True:
    urunleri_tara()
    try:
        tarayici.find_element(By.XPATH, "//a[normalize-space()='Sonraki']").click()
    except:
        print("Tüm sayfalar tarandı.")
        break

sleep(3)
tarayici.quit()
wb.save(excel_yolu)
