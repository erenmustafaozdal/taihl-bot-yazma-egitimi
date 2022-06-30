from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from datetime import datetime

aranacak_urun = input("Ne aramak istersiniz?: ")

# excel dosyası oluşturulur veya açılır
excel_path = "excel/amazon_urunleri.xlsx"
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    # aktif çalışma sayfası açılır
    ws = wb.active
else:
    wb = Workbook()
    # aktif çalışma sayfası açılır
    ws = wb.active
    # başlık satısı eklenir
    ws.append([
        "ZAMAN",
        "RESİM",
        "ID",
        "AD",
        "FİYAT",
        "ÇOK SATAN MI?",
        "FIRSAT MI?",
        "PRIME MI?",
        "DEĞERLENDİRME SAYISI",
        "DEĞERLENDİRME ORANI",
        "AZ STOK SAYISI"
    ])

tarayici = webdriver.Chrome(ChromeDriverManager().install())
tarayici.maximize_window()
tarayici.get("https://www.amazon.com.tr")

# çerezi kabul et
tarayici.find_element(By.ID, "sp-cc-accept").click()

tarayici.find_element(By.ID, "twotabsearchtextbox").send_keys(aranacak_urun)
tarayici.find_element(By.ID, "nav-search-submit-button").click()

# bekleme nesnesi oluşturulur
bekle = WebDriverWait(tarayici, 30)


def urunleri_tara():
    sleep(3)
    urunler = bekle.until(ec.visibility_of_all_elements_located(
        (By.XPATH,
         "//div[@data-component-type='s-search-result' and not(contains(@style, 'display: none'))]")
    ))
    # ürünler içince gez ve bilgilerini al
    for urun in urunler:
        # urun id'si
        id = urun.get_attribute("data-asin")
        print("id:", id)

        # adı
        ad = urun.find_element(By.TAG_NAME, "h2").text
        print("ad:", ad)

        # resim
        resim = f"resimler/{id}.png"
        urun.find_element(By.CLASS_NAME, "s-image").screenshot(resim)

        # fiyatı
        try:
            lira = urun.find_element(By.CLASS_NAME,
                                     "a-price-whole").text.replace(".", "")
            kurus = urun.find_element(By.CLASS_NAME, "a-price-fraction").text
            fiyat = float(lira + "." + kurus)
        except:
            fiyat = "Fiyat yok"
        print("fiyat:", fiyat)

        # çok satan mı?
        try:
            urun.find_element(By.ID, f"{id}-best-seller")
            cok_satan = True
        except:
            cok_satan = False
        print("Çok satan mı?:", cok_satan)

        # fırsat mı?
        try:
            urun.find_element(By.ID, f"BEST_DEAL_{id}")
            firsat = True
        except:
            firsat = False
        print("Fırsat mı?:", firsat)

        # prime mı?
        try:
            urun.find_element(By.CSS_SELECTOR, "span.s-prime")
            prime = True
        except:
            prime = False
        print("Prime mı?:", prime)

        # değerlendirmeler varsa al
        try:
            degerlendirme_orani_el = urun.find_element(By.XPATH,
                                                       ".//span[contains(@aria-label, '5 yıldız üzerinden')]")
            # değerlendirme sayısı
            degerlendirme_orani = float(
                degerlendirme_orani_el.get_attribute("aria-label").split()[
                    -1].replace(",", "."))
            # değerlendirme oranı
            degerlendirme_sayisi = int(
                degerlendirme_orani_el.find_element(By.XPATH,
                                                    ".//following-sibling::span").get_attribute(
                    "aria-label").replace(".", ""))
        except:
            degerlendirme_orani = "Değerlendirme yok"
            degerlendirme_sayisi = 0
        print("Değerlendirme Oranı:", degerlendirme_orani)
        print("Değerlendirme Sayısı:", degerlendirme_sayisi)

        # az stok sayısı
        try:
            az_stok_sayisi = int(urun.find_element(By.XPATH,
                                                   ".//span[contains(text(), 'Stokta sadece')]").text.split()[
                                     2])
        except:
            az_stok_sayisi = "Stok fazla"
        print("Az Stok Sayısı:", az_stok_sayisi)
        print("-" * 50)

        # excel'e yaz
        ws.append([
            datetime.now(),
            f'=HYPERLINK("{os.getcwd()}\\{resim}", "RESİM")',
            id,
            ad,
            fiyat,
            cok_satan,
            firsat,
            prime,
            degerlendirme_sayisi,
            degerlendirme_orani,
            az_stok_sayisi
        ])

    return urunler


# sonraki sayfaya geç
sayfa_sayaci = 1
while True:
    # ürünler alınır
    urunler = urunleri_tara()
    print()
    print("#"*50)
    print(f"{sayfa_sayaci}. sayfada {len(urunler)} ürün bulundu.")
    print("#"*50)
    print()

    try:
        tarayici.find_element(By.XPATH, "//a[contains(@class, 's-pagination-next')]").click()
    except:
        break

    sayfa_sayaci += 1


sleep(3)
tarayici.quit()
wb.save(excel_path)
