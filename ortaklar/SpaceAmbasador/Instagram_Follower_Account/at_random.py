from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from flags import kullanici_adi, sifre
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Create excel or open if any.
excel_path = "excel/Instagramfollowers.xlsx"
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active

    # Adds headers.
    ws.append(["Date", "Followers"])

driver = "C:\\Users\\GRAFIKLAB-18\\.wdm\\drivers\\geckodriver\\win64\\v0.31.0\\geckodriver.exe"

# Login to Instagram
tarayici = webdriver.Firefox(executable_path=driver)
tarayici.maximize_window()
tarayici.get("https://www.instagram.com/")
sleep(3)

tarayici.find_element(By.NAME, "username").click()
tarayici.find_element(By.NAME, "username").send_keys(kullanici_adi)

tarayici.find_element(By.NAME, "password").click()
tarayici.find_element(By.NAME, "password").send_keys(sifre)

sleep(2)

login_path = '//*[@id="loginForm"]/div/div[3]'

tarayici.find_element(By.XPATH, login_path).click()

sleep(5)


# Profile page reached.
profile_image = '//*[@id="react-root"]/section/nav/div[2]/div/div/div[3]/div/div[6]/div[1]'

tarayici.find_element(By.XPATH, profile_image).click()

sleep(3)


tarayici.find_element(By.CLASS_NAME, "-qQT3").click()
sleep(3)


# Followed button
tarayici.find_element(By.CSS_SELECTOR, 'body > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > '
                                       'div:nth-child(1) > div:nth-child(3) > div:nth-child(1) > div:nth-child(1) > '
                                       'div:nth-child(1) > div:nth-child(1) > section:nth-child(2) > main:nth-child('
                                       '2) > div:nth-child(1) > header:nth-child(1) > section:nth-child(2) > '
                                       'ul:nth-child(2) > li:nth-child(3) > a:nth-child(1) > div:nth-child(1').click()
sleep(3)


Followers = tarayici.find_elements(By.XPATH, "//div[contains(@class,'hwddc3l5')]//li//a/span")

for follower in Followers:
    print(f"Followers: {follower.text}")
    ws.append([
        datetime.now(),
        follower.text
        ])

# Saves excel file.
wb.save(excel_path)
 # Exits browser
tarayici.quit()
