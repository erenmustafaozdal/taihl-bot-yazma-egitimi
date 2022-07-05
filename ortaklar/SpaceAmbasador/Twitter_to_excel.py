from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from time import sleep
from settings import kullanici_adi, sifre, driver
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Create excel or open if any.
excel_yolu = "excel/Twitter_Trends.xlsx"
if os.path.exists(excel_yolu):
    wb = load_workbook(excel_yolu)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active

    # Adds headers.
    ws.append(["Date", "Trends"])

# Enters the browser.
tarayici = webdriver.Firefox(executable_path=driver)
tarayici.maximize_window()
tarayici.get("https://twitter.com/i/flow/login")

bekle = WebDriverWait(tarayici, 30)

# Logs in.
bekle.until(ec.visibility_of_element_located(
    (By.NAME, "text")
)).send_keys(kullanici_adi)
tarayici.find_element(By.XPATH, "//div[@role='button' and normalize-space()='İleri']").click()
bekle.until(ec.visibility_of_element_located(
    (By.NAME, "password")
)).send_keys(sifre)
tarayici.find_element(By.XPATH, "//div[@data-testid='LoginForm_Login_Button']").click()

# Enters the trends part.
bekle.until(ec.visibility_of_element_located(
    (By.XPATH, "//a[@href='/i/trends']")
)).click()
sleep(2)

# Prints trends to terminal and excel file.
Trends = tarayici.find_elements(By.XPATH,
                                '//div[@aria-label="Zaman Akışı: Gündemler"]/div/div/descendant::span['
                                '@class="r-18u37iz"]')

for trend in Trends:
    print(f"Trendler: {trend.text}")
    ws.append([
        datetime.now(),
        f'=HYPERLINK("https://twitter.com/hashtag/{trend.text.replace("#", "")}?src=hashtag_click", "{trend.text}")',

    ])

sleep(3)

# Takes a screenshot from trends page.
tarayici.save_screenshot('C:\\Users\\GRAFIKLAB-18\\Pictures\\PicturesTrends.png')
sleep(2)
wb.save(excel_yolu)

tarayici.quit()
