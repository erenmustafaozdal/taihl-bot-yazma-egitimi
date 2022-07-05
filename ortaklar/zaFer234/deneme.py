from openpyxl import Workbook, load_workbook
from time import sleep
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import os


driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install())
)
driver.get("https://www.igrus.com/guzel-sozler-kisa-anlamli-ve-bir-o-kadar-harika/?gclid=Cj0KCQjwn4qWBhCvARIsAFNAMiipOWH_G-D9pNpoKTCijFfr-Tk-qk6oBZvpVuhR4rgRMyFOH-NxyWAaArT2EALw_wcB")


excel_yolu = "liste.xlsx"
if os.path.exists(excel_yolu):
    wb = load_workbook(excel_yolu)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active


söz = driver.find_element(By.XPATH,"div[class='article-right-box'] ul li")

for sözler in 2000:
    yazı = söz.find_element(By.ID, "post-42409").text
    print(yazı)
    ws.append(yazı)
